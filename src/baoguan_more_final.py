from openpyxl import load_workbook
import os
import shutil 
import re 
from datetime import datetime
import glob

# 第一段程序
def part1():
    print("执行第一部分代码...")
    #基于这个openpyxl库不支持计算，在第一次计算后，再次打开不能加载公式的值
    #基于程序需要系统在没有安装原生Excel的情况下兼容WPS运行
    #基于程序需要跨windows和MAC平台运行
    #采用openpyxl，加断点执行加载中间数据的方式

    #第一步：计算立方运算计算模型
    #初始化预录表
    #提示用户输入预录表路径，并且给予一些用户提示
    global yulu_path
    yulu_path = input("请输入文件的绝对路径：").strip()
    #yulu_path = r"D:\VScode\Python_File\2025.03.06-宠物尿垫-523box-8910件\20250306宠物尿垫报关退税预录表.xlsx"
    yulu_wb = load_workbook(yulu_path, data_only=True)
    yulu_sheet = yulu_wb.active
    #初始化立方运费计算模型
    date_str = datetime.now().strftime("%Y%m%d")
    lifang_path_name = f"step1-立方运费计算模型.xlsx"
    lifang_path = os.path.join(os.path.dirname(yulu_path), lifang_path_name)
    #lifang_path = r"D:\VScode\Python_File\2025.03.06-宠物尿垫-523box-8910件\step1-立方运费计算模型-20250306-更新.xlsx"
    lifang_wb = load_workbook(lifang_path)#可以保存赋值和公式
    lifang_sheet = lifang_wb.active
    #计算过程
    lifang_sheet["B6"].value = yulu_sheet["F10"].value
    lifang_sheet["B7"] = yulu_sheet["E10"].value
    lifang_sheet["B24"] = yulu_sheet["E10"].value
    lifang_sheet["B9"] = yulu_sheet["I10"].value
    lifang_sheet["B10"] = yulu_sheet["J10"].value
    lifang_sheet["B11"] = yulu_sheet["D10"].value
    lifang_sheet["e22"] = yulu_sheet["G10"].value

    for n,m in zip(range(5), ["f","i","l","o","r"]):
        lifang_sheet[f"{m}4"] = yulu_sheet[f"D{n+2}"].value[0:4]
        lifang_sheet[f"{m}7"] = yulu_sheet[f"E{n+2}"].value
        lifang_sheet[f"{m}6"] = yulu_sheet[f"F{n+2}"].value
    lifang_wb.save(lifang_path)
    #计算完成后打开立方运费计算模型
    #第一次 data_only=True 获取计算后的值，但不能修改 Excel
    #第二次 不加 data_only，以允许修改并保存 Excel。    

# 第二段程序
def part2():
    print("\n执行第二部分代码...")
    global yulu_path
    #yulu_path = r"D:\VScode\Python_File\2025.03.06-宠物尿垫-523box-8910件\20250306宠物尿垫报关退税预录表.xlsx"
    yulu_wb = load_workbook(yulu_path, data_only=True)
    yulu_sheet = yulu_wb.active
    #初始化保存公式计算后的立方计算模型
    date_str = datetime.now().strftime("%Y%m%d")
    lifang_path_name = f"step1-立方运费计算模型.xlsx"
    lifang_path = os.path.join(os.path.dirname(yulu_path), lifang_path_name)
    lifang_wb = load_workbook(lifang_path, data_only=True)
    lifang_sheet = lifang_wb.active
    #输出这次计算每片商品报关值验证是否输出正确
    print(lifang_sheet["b34"].value)

    #计算报关表
    for n,m in zip(range(5), ["f","i","l","o","r"]):
        #新建报关子文件夹
        baoguan_folder = yulu_sheet.cell(row=n+2, column=9).value
        baoguan_path = os.path.join(os.path.dirname(yulu_path), baoguan_folder)
        os.mkdir(baoguan_path)
        #新建报关子表
        baoguan_sheet_name = f"{baoguan_folder}-报关退税.xlsx"
        baoguan_sheet_path = os.path.join(baoguan_path,baoguan_sheet_name)
        baoguan_sheet_model_path = os.path.join(os.path.dirname(yulu_path), "报关表模版.xlsx")
        # 执行复制操作（含异常处理）
        try:
            shutil.copy(baoguan_sheet_model_path,  baoguan_sheet_path)
            print(f"文件已复制并重命名为：{baoguan_sheet_path}")
        except FileNotFoundError:   
            print("源文件不存在或路径错误！")
        except PermissionError:
            print("权限不足，无法写入目标路径！")
        #加载报关子表
        baoguan_sheet_wb = load_workbook(baoguan_sheet_path)
        #初始化报关子表sheet
        fapiao_sheet = baoguan_sheet_wb["发票"]
        xiangdan_sheet = baoguan_sheet_wb["箱单"]
        weituo_sheet = baoguan_sheet_wb["委托书"]
        baoguandan_sheet = baoguan_sheet_wb["报关单"]
        yaosu_sheet = baoguan_sheet_wb["申报要素"]
        #发票
        fapiao_sheet["B4"] = yulu_sheet[f"D{n+2}"].value
        fapiao_sheet["E4"] = yulu_sheet[f"B{n+2}"].value
        
        fapiao_sheet["E7"] = "03.28.2025"
        
        #名字做截取处理:委托书、发票、申报要素的名称都在这里填
        print(yulu_sheet["A10"].value)
        pattern = r"[（(](.*?)[）)]"  # 匹配任意类型括号 
        match = re.search(pattern, yulu_sheet["A10"].value)
        if match:
            fapiao_sheet["A10"] = match.group(1).strip()               # 括号内内容：pet pee pad 
            yaosu_sheet["B6"] = match.group(1).strip()
            weituo_sheet["H21"] = re.sub(pattern,  "", yulu_sheet["A10"].value).strip()     # 括号外内容：宠物尿垫 
        else:
            weituo_sheet["H21"] = yulu_sheet["A10"].value.strip() 
            fapiao_sheet["A10"] = ""
            yaosu_sheet["B6"] = ""

        fapiao_sheet["C10"] = yulu_sheet[f"E{n+2}"].value
        fapiao_sheet["F10"] = lifang_sheet["B34"].value
        #箱单
        xiangdan_sheet["C10"] = yulu_sheet[f"F{n+2}"].value
        xiangdan_sheet["D10"] = lifang_sheet[f"{m}10"].value
        #委托单
        weituo_sheet["O20"] = lifang_sheet[f"{m}9"].value
        #报关单
        baoguandan_sheet["K12"] = lifang_sheet["F13"].value
        #报关要素
        
        yaosu_sheet["B8"] = yulu_sheet["B10"].value    

        #保存报关子表
        baoguan_sheet_wb.save(baoguan_sheet_path)

# 主程序流程
if __name__ == "__main__":
    part1()
    
    while True:
        choice = input("\n第一部分执行完毕，是否继续？(Y/N): ").strip().upper()
        if choice == "Y":
            part2()
            break
        elif choice == "N":
            print("操作已终止")
            break
        else:
            print("无效输入，请按 Y 或 N 键确认")